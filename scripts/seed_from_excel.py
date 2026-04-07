"""
One-time seed script: reads strategy performance data from
'HIP3 - Strategy Performance.xlsx' and populates the vault tables
so the Vault dashboard page can display historical numbers.

Targets:
  - vault_snapshots (total vault equity + APR per date)
  - vault_strategy_snapshots (per-strategy equity + APR per date)
  - vault_cashflows (DEPOSIT/WITHDRAW events)

Pre-requisite: run `vault.py sync-registry` first to populate vault_strategies.

Usage:
    source .arbit_env && .venv/bin/python scripts/seed_from_excel.py
"""

from __future__ import annotations

import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "HIP3 - Strategy Performance.xlsx"
DB_PATH = PROJECT_ROOT / "tracking" / "db" / "arbit_v3.db"

# Strategy IDs must match config/strategies.json
STRATEGY_MAP = {
    "Lending": "lending",
    "Delta Neutral": "delta_neutral",
    "Depeg": "depeg",
}

SEED_META = '{"source": "excel_seed"}'


def dt_to_ms(dt: datetime) -> int:
    """Convert a naive datetime (assumed UTC) to epoch milliseconds."""
    return int(dt.replace(tzinfo=timezone.utc).timestamp() * 1000)


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------
def read_overview(wb: openpyxl.Workbook) -> list[dict]:
    """Parse the Overview sheet into a list of date-point dicts."""
    ws = wb["Overview"]
    rows = list(ws.iter_rows(min_row=2, max_col=7, values_only=True))

    date_points = []
    current_date = None
    group: dict = {}

    for row in rows:
        date_val, strategy, equity, apr, weight, total_apr, _total_eq = row

        if date_val is not None and isinstance(date_val, datetime):
            if group and group.get("strategies"):
                date_points.append(group)
            current_date = date_val
            group = {"date": current_date, "strategies": {}, "total_apr": None}

        if strategy is None or equity is None:
            continue

        name = strategy.strip()
        group["strategies"][name] = {
            "equity": equity,
            "apr": apr,
            "weight": weight,
        }
        if total_apr is not None:
            group["total_apr"] = total_apr

    if group and group.get("strategies"):
        date_points.append(group)

    return date_points


def read_cashflows(wb: openpyxl.Workbook) -> list[dict]:
    """Read cash flow events from Lending and Delta Neutral sheets."""
    cfs = []

    # Lending cashflows
    ws = wb["Lending"]
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        date_val, _eq, _apr, cf = row
        if cf is not None and date_val is not None:
            cfs.append({
                "date": date_val,
                "amount": cf,
                "strategy_id": "lending",
            })

    # Delta Neutral cashflows
    ws = wb["Delta Neutral"]
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        date_val, _eq, _pnl, _funding, _apr, cf = row
        if cf is not None and date_val is not None:
            cfs.append({
                "date": date_val,
                "amount": cf,
                "strategy_id": "delta_neutral",
            })

    return cfs


# ---------------------------------------------------------------------------
# Seeding logic
# ---------------------------------------------------------------------------
def seed(con: sqlite3.Connection) -> dict[str, int]:
    """Seed the vault tables. Returns counts of rows inserted per table."""
    counts = {
        "vault_snapshots": 0,
        "vault_strategy_snapshots": 0,
        "vault_cashflows": 0,
    }

    # Verify vault_strategies is populated
    strat_count = con.execute("SELECT COUNT(*) FROM vault_strategies").fetchone()[0]
    if strat_count == 0:
        print("ERROR: vault_strategies is empty. Run: vault.py sync-registry")
        return counts

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    date_points = read_overview(wb)
    cashflows = read_cashflows(wb)

    # ---- 1. Clear previous data (seed + any live snapshots) ----
    con.execute("DELETE FROM vault_snapshots")
    con.execute("DELETE FROM vault_strategy_snapshots")
    con.execute("DELETE FROM vault_cashflows WHERE meta_json = ?", (SEED_META,))

    # ---- 2. Compute net deposits for each date point ----
    net_deposits_cum = 0.0
    net_deposits_by_date: dict[str, float] = {}
    # Sort cashflows by date
    sorted_cfs = sorted(cashflows, key=lambda c: c["date"])
    for cf in sorted_cfs:
        net_deposits_cum += cf["amount"]
        net_deposits_by_date[cf["date"].strftime("%Y-%m-%d")] = net_deposits_cum

    # ---- 3. Seed vault_strategy_snapshots and vault_snapshots ----
    for dp in date_points:
        ts = dt_to_ms(dp["date"])
        strats = dp["strategies"]
        total_equity = 0.0
        weights: dict[str, float] = {}

        # Per-strategy snapshots
        for display_name, strategy_id in STRATEGY_MAP.items():
            s = strats.get(display_name)
            if s is None:
                continue

            equity = s["equity"] or 0.0
            raw_apr = s.get("apr")  # per-strategy APR (since inception)
            apr = raw_apr * 100 if raw_apr else None  # Excel stores decimal; DB uses percent
            total_equity += equity

            # Delete any existing snapshot for same strategy + day bucket
            day_bucket = ts // 86400000
            con.execute(
                "DELETE FROM vault_strategy_snapshots "
                "WHERE strategy_id = ? AND CAST(ts / 86400000 AS INTEGER) = ?",
                (strategy_id, day_bucket),
            )

            con.execute(
                "INSERT INTO vault_strategy_snapshots "
                "(strategy_id, ts, equity_usd, equity_breakdown_json, "
                "apr_since_inception, apr_30d, apr_7d, meta_json) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    strategy_id, ts, equity, None,
                    apr,  # apr_since_inception
                    apr,  # apr_30d (use same value as approximation)
                    apr,  # apr_7d (use same value as approximation)
                    SEED_META,
                ),
            )
            counts["vault_strategy_snapshots"] += 1

        # Strategy weights
        if total_equity > 0:
            for display_name, strategy_id in STRATEGY_MAP.items():
                s = strats.get(display_name)
                if s and s["equity"]:
                    weights[strategy_id] = round(s["equity"] / total_equity * 100, 2)

        # Vault-level snapshot
        raw_total_apr = dp.get("total_apr")
        total_apr = raw_total_apr * 100 if raw_total_apr else None  # Excel decimal → percent
        date_key = dp["date"].strftime("%Y-%m-%d")

        # Find net deposits up to this date
        net_deps = 0.0
        for cf_date_str, cum_val in net_deposits_by_date.items():
            if cf_date_str <= date_key:
                net_deps = cum_val

        day_bucket = ts // 86400000
        con.execute(
            "DELETE FROM vault_snapshots WHERE CAST(ts / 86400000 AS INTEGER) = ?",
            (day_bucket,),
        )

        con.execute(
            "INSERT INTO vault_snapshots "
            "(ts, total_equity_usd, strategy_weights_json, "
            "total_apr, apr_30d, apr_7d, net_deposits_alltime, meta_json) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                ts, total_equity,
                json.dumps(weights, separators=(",", ":")),
                total_apr,
                total_apr,  # apr_30d approximation
                total_apr,  # apr_7d approximation
                net_deps,
                SEED_META,
            ),
        )
        counts["vault_snapshots"] += 1

    # ---- 4. Seed vault_cashflows ----
    now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
    for cf in cashflows:
        ts = dt_to_ms(cf["date"])
        amount = cf["amount"]
        cf_type = "DEPOSIT" if amount > 0 else "WITHDRAW"
        signed = amount  # WITHDRAW already negative from Excel

        con.execute(
            "INSERT INTO vault_cashflows "
            "(ts, cf_type, amount, strategy_id, currency, description, meta_json, created_at_ms) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                ts, cf_type, signed, cf["strategy_id"],
                "USDC", "Seeded from Excel", SEED_META, now_ms,
            ),
        )
        counts["vault_cashflows"] += 1

    con.commit()
    return counts


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return
    if not DB_PATH.exists():
        print(f"ERROR: Database not found at {DB_PATH}")
        return

    con = sqlite3.connect(str(DB_PATH))
    con.execute("PRAGMA journal_mode=WAL")
    con.row_factory = sqlite3.Row

    try:
        counts = seed(con)
    except Exception:
        con.rollback()
        raise
    finally:
        con.close()

    print("Seed complete. Rows inserted:")
    for table, count in counts.items():
        print(f"  {table}: {count}")
    print(f"  Total: {sum(counts.values())}")


if __name__ == "__main__":
    main()
